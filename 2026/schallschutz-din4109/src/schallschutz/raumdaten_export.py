"""Raum-zentrischer Excel-Export der aus dem Grundriss erarbeiteten Raumdaten.

Format nach Nutzer-Vorlage ("Vorlauge Dropdown.xlsx"), best-practice umgesetzt:
- jeder Raum als Ueberpunkt (Bezeichnung, Flaeche m2, Volumen m3)
- darunter die einzelnen Waende, per Excel-Gruppierung EIN-/AUSKLAPPBAR ("aufklickbar")
- je Wand: Bezug (anliegender Raum, Aussenwand = "A"), Material, Dicke, Hoehe, Laenge, Fensterflaeche
- echte Dropdowns (Datenvalidierung): Bezug = Raumliste + "A"; Material = Materialkatalog
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .engine import _CURVE_CATS
from .knowledge import materials, resolve_material
from .models import Bauteiltyp, Projekt

_HFILL = PatternFill("solid", fgColor="33415C")
_HFONT = Font(bold=True, color="FFFFFF")
_ROOMFILL = PatternFill("solid", fgColor="DCE6F1")
_ROOMFONT = Font(bold=True, size=11)
_thin = Side(style="thin", color="C0C7D0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center")

# Spalten der Raumdaten-Tabelle
_COLS = ["Pos.", "Bezeichnung / Bezug", "Material", "Dicke [cm]", "Höhe [m]",
         "Länge [m]", "Fenster [m²]", "Fläche [m²]", "Volumen [m³]"]


def _kernschicht(bt):
    """Liefert (material_key, dicke_mm) der tragenden Schicht einer Wand (groesste Kurven-Masse)."""
    best = None
    for s in bt.schichten:
        key, mat = resolve_material(s.material)
        if mat and mat["kategorie"] in _CURVE_CATS:
            masse = (s.dicke_mm / 1000.0) * (s.rohdichte_override or mat["rohdichte"])
            if best is None or masse > best[2]:
                best = (key, s.dicke_mm, masse)
    if best:
        return best[0], best[1]
    # Fallback: erste Schicht
    if bt.schichten:
        key, _ = resolve_material(bt.schichten[0].material)
        return key or bt.schichten[0].material, bt.schichten[0].dicke_mm
    return None, None


def _waende_je_raum(projekt: Projekt) -> dict[str, list[dict]]:
    """Ordnet jede Wand ihren angrenzenden Raeumen zu (raum-individuell, Wand erscheint bei beiden)."""
    namen = {r.id: r.name for r in projekt.raeume}
    zuordnung: dict[str, list[dict]] = {r.id: [] for r in projekt.raeume}
    nicht_zugeordnet: list[dict] = []
    for bt in projekt.bauteile:
        if bt.typ != Bauteiltyp.WAND:
            continue
        mat, dicke = _kernschicht(bt)
        seiten = [(bt.raum_a, bt.raum_b), (bt.raum_b, bt.raum_a)]
        traf = False
        eintrag = {
            "wand_id": bt.id, "material": mat,
            "dicke_cm": round(dicke / 10, 1) if dicke else None,
            "hoehe_m": bt.hoehe_m, "laenge_m": bt.laenge_m, "fenster_m2": bt.fenster_flaeche_m2,
            "din_rolle": bt.din_rolle, "bemerkung": bt.bemerkung,
        }
        for eigen, nachbar in seiten:
            if eigen in zuordnung:
                bezug = (namen.get(nachbar, nachbar) if nachbar else "A")  # unbek. ID -> ID, sonst A=Aussenwand
                zuordnung[eigen].append({**eintrag, "bezug": bezug or "A"})
                traf = True
        if not traf:
            nicht_zugeordnet.append({**eintrag, "bezug": "?"})
    return zuordnung, nicht_zugeordnet


def export_raumdaten_excel(projekt: Projekt, pfad: str | Path) -> Path:
    pfad = Path(pfad)
    wb = Workbook()
    ws = wb.active
    ws.title = "Raumdaten"
    ws.sheet_properties.outlinePr.summaryBelow = False  # Raum-Kopf steht UEBER seinen Waenden

    # Dropdown-Quellen auf separatem Blatt
    listen = wb.create_sheet("Listen")
    listen["A1"] = "Raeume/Bezug"
    listen["C1"] = "Materialien"
    raum_namen = [r.name for r in projekt.raeume] + ["A"]
    for i, n in enumerate(raum_namen, start=2):
        listen.cell(i, 1, n)
    mat_keys = list(materials().keys())
    for i, m in enumerate(mat_keys, start=2):
        listen.cell(i, 3, m)
    listen.sheet_state = "hidden"

    dv_bezug = DataValidation(type="list", formula1=f"=Listen!$A$2:$A${len(raum_namen)+1}", allow_blank=True)
    dv_mat = DataValidation(type="list", formula1=f"=Listen!$C$2:$C${len(mat_keys)+1}", allow_blank=True)
    ws.add_data_validation(dv_bezug)
    ws.add_data_validation(dv_mat)

    # Titel + Kopfzeile
    ws["A1"] = f"Raumdaten — {projekt.name}"
    ws["A1"].font = Font(bold=True, size=14)
    if projekt.bauvorhaben:
        ws["A2"] = projekt.bauvorhaben
        ws["A2"].font = Font(italic=True, color="57606A")
    hdr = 4
    for j, c in enumerate(_COLS, start=1):
        cell = ws.cell(hdr, j, c)
        cell.font = _HFONT
        cell.fill = _HFILL
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.freeze_panes = f"A{hdr+1}"

    zuordnung, nicht_zugeordnet = _waende_je_raum(projekt)
    r = hdr + 1

    def schreibe_raum(raum_id, name, flaeche, volumen, waende):
        nonlocal r
        # Raum-Kopfzeile (Outline-Level 0)
        ws.cell(r, 1, raum_id).font = _ROOMFONT
        ws.cell(r, 2, name).font = _ROOMFONT
        ws.cell(r, 8, flaeche)
        ws.cell(r, 9, volumen)
        for j in range(1, 10):
            ws.cell(r, j).fill = _ROOMFILL
            ws.cell(r, j).border = _BORDER
        r += 1
        # Wandzeilen (Outline-Level 1, einklappbar)
        liste = waende or [{}]  # mind. eine Leerzeile zum Ausfuellen
        for k, w in enumerate(liste, start=1):
            ws.cell(r, 1, f"Wand {k}")
            ws.cell(r, 2, w.get("bezug", ""))
            ws.cell(r, 3, w.get("material", ""))
            ws.cell(r, 4, w.get("dicke_cm", ""))
            ws.cell(r, 5, w.get("hoehe_m", ""))
            ws.cell(r, 6, w.get("laenge_m", ""))
            ws.cell(r, 7, w.get("fenster_m2", ""))
            for j in range(1, 10):
                ws.cell(r, j).border = _BORDER
            dv_bezug.add(ws.cell(r, 2))
            dv_mat.add(ws.cell(r, 3))
            ws.row_dimensions[r].outlineLevel = 1
            r += 1

    for raum in projekt.raeume:
        vol = round(raum.flaeche_m2 * raum.hoehe_m, 1) if (raum.flaeche_m2 and raum.hoehe_m) else None
        schreibe_raum(raum.id, raum.name, raum.flaeche_m2, vol, zuordnung.get(raum.id, []))

    if nicht_zugeordnet:
        schreibe_raum("(nicht zugeordnet)", "Wände ohne eindeutige Raumzuordnung", None, None, nicht_zugeordnet)

    # Spaltenbreiten
    for j, w in enumerate([16, 30, 20, 10, 9, 9, 12, 12, 13], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    wb.save(pfad)
    return pfad
