"""Export der Nachweisdaten als Eingabe fuer Schallschutz-Berechnungssoftware (Excel/CSV/JSON).

Liefert eine strukturierte Bauteilliste mit Aufbau, m', erf./vorh. R'w und L'n,w -
das typische Eingabeformat, das Ingenieure sonst manuell in Tools wie BASTIAN/Daemmwerk uebertragen.
"""
from __future__ import annotations

import json
from pathlib import Path

from .nachweis import NachweisErgebnis

_SPALTEN = [
    "Bauteil-ID", "Bauteiltyp", "DIN-Rolle", "Bezeichnung", "DIN-Tabelle",
    "Aufbau", "m_flaechen_kg_m2", "Massekurve",
    "erf_Rw_dB", "vorh_Rw_dB", "Luftschall_OK",
    "zul_Lnw_dB", "vorh_Lnw_dB", "Trittschall_OK",
    "delta_Rw_Vorsatz", "delta_Lw_Estrich", "Status",
]


def _zeilen_dicts(res: NachweisErgebnis) -> list[dict]:
    rows = []
    for z in res.zeilen:
        e = z.ergebnis
        aufbau = " | ".join(
            f"{s.material} {s.dicke_mm:.0f}mm"
            + (f"@{s.rohdichte:.0f}" if s.rohdichte else "")
            for s in e.schichten
        )
        rows.append({
            "Bauteil-ID": z.bauteil_id, "Bauteiltyp": z.typ, "DIN-Rolle": z.din_rolle or "",
            "Bezeichnung": z.bezeichnung, "DIN-Tabelle": z.tabelle_zeile,
            "Aufbau": aufbau, "m_flaechen_kg_m2": e.masse_kg_m2 or "", "Massekurve": e.massekurve or "",
            "erf_Rw_dB": z.erf_rw if z.erf_rw is not None else "",
            "vorh_Rw_dB": z.vorh_rw if z.vorh_rw is not None else "",
            "Luftschall_OK": "" if z.rw_ok is None else ("ja" if z.rw_ok else "nein"),
            "zul_Lnw_dB": z.zul_lnw if z.zul_lnw is not None else "",
            "vorh_Lnw_dB": z.vorh_lnw if z.vorh_lnw is not None else "",
            "Trittschall_OK": "" if z.lnw_ok is None else ("ja" if z.lnw_ok else "nein"),
            "delta_Rw_Vorsatz": next((bt.delta_rw_vorsatz for bt in res.projekt.bauteile if bt.id == z.bauteil_id), ""),
            "delta_Lw_Estrich": next((bt.delta_lw for bt in res.projekt.bauteile if bt.id == z.bauteil_id), ""),
            "Status": {"gruen": "erfuellt", "rot": "nicht erfuellt", "offen": "offen"}[z.status],
        })
    return rows


def export_excel(res: NachweisErgebnis, pfad: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    pfad = Path(pfad)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bauteilkatalog"
    ws.append(_SPALTEN)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDE3EA")
    fill = {"erfuellt": "D7F0D9", "nicht erfuellt": "F7D4D7", "offen": "FBEFC9"}
    for row in _zeilen_dicts(res):
        ws.append([row[s] for s in _SPALTEN])
        f = fill.get(row["Status"])
        if f:
            ws.cell(ws.max_row, len(_SPALTEN)).fill = PatternFill("solid", fgColor=f)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)
    wb.save(pfad)
    return pfad


def export_json(res: NachweisErgebnis, pfad: str | Path) -> Path:
    pfad = Path(pfad)
    payload = {
        "projekt": res.projekt.name,
        "bauvorhaben": res.projekt.bauvorhaben,
        "gebaeudetyp": res.projekt.gebaeude.typ.value,
        "gesamt_status": res.gesamt_status,
        "bauteile": _zeilen_dicts(res),
    }
    pfad.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return pfad
