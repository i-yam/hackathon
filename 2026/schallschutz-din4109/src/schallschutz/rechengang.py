"""Transparenter Rechengang als Excel-Arbeitsmappe mit ECHTEN Excel-Formeln.

Vier Blaetter:
  1 Uebersicht        Projektkopf + Gesamtergebnis + Zusammenfassung
  2 Eingabeparameter  alle Inputs sauber gelistet (Schichten, Rohdichten, Korrekturwerte)
  3 Formeln & Normwerte  verwendete DIN-Formeln + Anforderungstabellen
  4 Rechengang        je Bauteil Schritt fuer Schritt mit LIVE-Formeln (m' -> R_w -> R'w -> Nachweis)

Die Zahlen sind keine Konstanten: jede Zelle enthaelt die Formel mit Zellbezuegen,
d.h. man kann eine Dicke aendern und alles rechnet in Excel neu. Voll nachvollziehbar.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .knowledge import materials, requirements
from .nachweis import NachweisErgebnis, NachweisZeile

# -- Stile -------------------------------------------------------------------
_TITLE = Font(bold=True, size=15, color="1c2128")
_H = Font(bold=True, size=11, color="FFFFFF")
_HFILL = PatternFill("solid", fgColor="33415c")
_BLOCK = Font(bold=True, size=11, color="1c2128")
_BLOCKFILL = PatternFill("solid", fgColor="DDE3EA")
_LBL = Font(bold=True)
_MONO = Font(name="Consolas", size=10)
_GREEN = PatternFill("solid", fgColor="D7F0D9")
_RED = PatternFill("solid", fgColor="F7D4D7")
_YELLOW = PatternFill("solid", fgColor="FBEFC9")
_RESULT = Font(bold=True, size=11)
_thin = Side(style="thin", color="B0B7C0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_RIGHT = Alignment(horizontal="right")
_WRAP = Alignment(wrap_text=True, vertical="top")

_STATUSFILL = {"gruen": _GREEN, "rot": _RED, "offen": _YELLOW}
_STATUSTXT = {"gruen": "erfuellt", "rot": "NICHT erfuellt", "offen": "offen"}


def _hdr(ws, row, labels, start=1):
    for i, t in enumerate(labels):
        c = ws.cell(row, start + i, t)
        c.font = _H
        c.fill = _HFILL
        c.border = _BORDER
    return row + 1


def _rw_formel(massekurve: str, mcell: str) -> tuple[str, str]:
    """Liefert (Excel-Formel, Klartext-Quelle) fuer R_w aus m' je Massekurve."""
    if massekurve in ("beton", "ks", "ziegel", "verfuellstein"):
        return f"=30.9*LOG10({mcell})-22.2", "DIN 4109-32 Gl.(13): 30,9*lg(m')-22,2"
    if massekurve == "leichtbeton":
        return f"=30.9*LOG10({mcell})-20.2", "DIN 4109-32 Gl.(14): 30,9*lg(m')-20,2"
    if massekurve == "porenbeton":
        return (f"=IF({mcell}<150,32.6*LOG10({mcell})-22.5,26.1*LOG10({mcell})-8.4)",
                "DIN 4109-32 Gl.(15/16): <150: 32,6*lg-22,5 ; >=150: 26,1*lg-8,4")
    return "", ""


# ============================================================ Blatt 1: Uebersicht
def _blatt_uebersicht(wb, res: NachweisErgebnis):
    ws = wb.active
    ws.title = "1 Uebersicht"
    p = res.projekt
    ws["A1"] = "Schallschutznachweis nach DIN 4109"
    ws["A1"].font = _TITLE
    meta = [
        ("Projekt", p.name), ("Bauvorhaben", p.bauvorhaben or "-"),
        ("Gebaeudetyp", p.gebaeude.typ.value), ("Datenquelle (Plan)", p.quelle_plan or "-"),
        ("Bearbeiter", p.bearbeiter or "-"),
    ]
    r = 3
    for k, v in meta:
        ws.cell(r, 1, k).font = _LBL
        ws.cell(r, 2, v)
        r += 1
    r += 1
    ws.cell(r, 1, "Gesamtergebnis:").font = _RESULT
    g = ws.cell(r, 2, _STATUSTXT[res.gesamt_status].upper())
    g.font = _RESULT
    g.fill = _STATUSFILL[res.gesamt_status]
    ws.cell(r, 3, f"{res.anzahl_ok} erfuellt | {res.anzahl_rot} nicht erfuellt | {res.anzahl_offen} offen")
    r += 2

    head = ["ID", "Bauteil", "erf. R'w", "vorh. R'w", "zul. L'n,w", "vorh. L'n,w", "Ergebnis"]
    r = _hdr(ws, r, head)
    for z in res.zeilen:
        ws.cell(r, 1, z.bauteil_id)
        ws.cell(r, 2, z.bezeichnung)
        ws.cell(r, 3, z.erf_rw if z.erf_rw is not None else "-").alignment = _RIGHT
        ws.cell(r, 4, z.vorh_rw if z.vorh_rw is not None else "-").alignment = _RIGHT
        ws.cell(r, 5, z.zul_lnw if z.zul_lnw is not None else "-").alignment = _RIGHT
        ws.cell(r, 6, z.vorh_lnw if z.vorh_lnw is not None else "-").alignment = _RIGHT
        sc = ws.cell(r, 7, _STATUSTXT[z.status])
        sc.fill = _STATUSFILL[z.status]
        sc.font = _LBL
        for c in range(1, 8):
            ws.cell(r, c).border = _BORDER
        r += 1
    _widths(ws, {1: 8, 2: 46, 3: 10, 4: 10, 5: 11, 6: 11, 7: 14})


# ======================================================= Blatt 2: Eingabeparameter
def _blatt_eingabe(wb, res: NachweisErgebnis):
    ws = wb.create_sheet("2 Eingabeparameter")
    ws["A1"] = "Eingabeparameter (aus Plan extrahiert / geprueft)"
    ws["A1"].font = _TITLE

    r = 3
    ws.cell(r, 1, "A) Bauteilaufbauten").font = _BLOCK
    ws.cell(r, 1).fill = _BLOCKFILL
    r += 1
    r = _hdr(ws, r, ["Bauteil", "DIN-Rolle", "Schicht", "Material", "d [mm]", "rho [kg/m3]", "m'_i [kg/m2]", "zaehlt zu m'"])
    for z in res.zeilen:
        e = z.ergebnis
        if not e.schichten:
            ws.cell(r, 1, z.bauteil_id)
            ws.cell(r, 2, z.din_rolle or "-")
            ws.cell(r, 4, f"(Element-Rw = {z.ergebnis.rw_konstruktion} dB, kein Schichtaufbau)")
            r += 1
            continue
        for i, s in enumerate(e.schichten):
            ws.cell(r, 1, z.bauteil_id if i == 0 else "")
            ws.cell(r, 2, (z.din_rolle or "-") if i == 0 else "")
            ws.cell(r, 3, i + 1)
            ws.cell(r, 4, s.material + (f"  [{s.erkannt_als}]" if s.erkannt_als else ""))
            ws.cell(r, 5, s.dicke_mm).alignment = _RIGHT
            ws.cell(r, 6, s.rohdichte if s.rohdichte else "-").alignment = _RIGHT
            ws.cell(r, 7, s.masse_kg_m2 if s.masse_kg_m2 else "-").alignment = _RIGHT
            ws.cell(r, 8, "ja" if s.in_masse else "nein (schwimmend)").alignment = _RIGHT
            r += 1
    r += 1
    ws.cell(r, 1, "B) Zusatzwerte je Bauteil").font = _BLOCK
    ws.cell(r, 1).fill = _BLOCKFILL
    r += 1
    r = _hdr(ws, r, ["Bauteil", "dRw Vorsatz [dB]", "dLw Estrich/Belag [dB]", "K Flanke Luft [dB]", "K_T Tritt [dB]"])
    for z in res.zeilen:
        e = z.ergebnis
        ws.cell(r, 1, z.bauteil_id)
        bt = next((b for b in res.projekt.bauteile if b.id == z.bauteil_id), None)
        ws.cell(r, 2, bt.delta_rw_vorsatz if bt else "-").alignment = _RIGHT
        ws.cell(r, 3, (e.dlw_verwendet if e.dlw_verwendet is not None else (bt.delta_lw if bt else 0))).alignment = _RIGHT
        ws.cell(r, 4, e.k_luft if e.k_luft is not None else "-").alignment = _RIGHT
        ws.cell(r, 5, e.k_tritt if e.k_tritt is not None else "-").alignment = _RIGHT
        r += 1
    r += 1
    ws.cell(r, 1, "C) Verwendete Materialkennwerte (Rohdichte-Rechenwerte)").font = _BLOCK
    ws.cell(r, 1).fill = _BLOCKFILL
    r += 1
    r = _hdr(ws, r, ["Material-Schluessel", "Kategorie/Massekurve", "rho [kg/m3]", "Quelle"])
    used = {s.erkannt_als for z in res.zeilen for s in z.ergebnis.schichten if s.erkannt_als}
    mats = materials()
    for key in sorted(x for x in used if x in mats):
        m = mats[key]
        ws.cell(r, 1, key)
        ws.cell(r, 2, m["kategorie"])
        ws.cell(r, 3, m["rohdichte"]).alignment = _RIGHT
        ws.cell(r, 4, m.get("quelle", ""))
        r += 1
    _widths(ws, {1: 16, 2: 26, 3: 22, 4: 18, 5: 16, 6: 14, 7: 14, 8: 18})


# ===================================================== Blatt 3: Formeln & Normwerte
def _blatt_formeln(wb, res: NachweisErgebnis):
    ws = wb.create_sheet("3 Formeln & Normwerte")
    ws["A1"] = "Formeln & normative Grundlagen"
    ws["A1"].font = _TITLE
    formeln = [
        ("Flaechenbezogene Masse", "m' = Sum( d_i * rho_i )", "DIN 4109-32:2016, Gl.(3); d in m, rho in kg/m3"),
        ("R_w Beton/KS/Ziegel", "R_w = 30,9 * lg(m') - 22,2", "DIN 4109-32 Gl.(13), 65..720 kg/m2"),
        ("R_w Leichtbeton", "R_w = 30,9 * lg(m') - 20,2", "DIN 4109-32 Gl.(14), 140..480 kg/m2"),
        ("R_w Porenbeton", "<150: 32,6*lg(m')-22,5 ; >=150: 26,1*lg(m')-8,4", "DIN 4109-32 Gl.(15/16)"),
        ("Vorsatzschale", "R_w,ges = R_w + dR_w", "DIN 4109-34 (Luftschallverbesserung)"),
        ("Eingebaut Luftschall", "R'_w = R_w,ges - K_Flanke", "DIN 4109-2 (Flankenuebertragung, hier vereinfacht)"),
        ("Norm-Trittschallpegel", "L_n,eq,0,w = 164 - 35 * lg(m')", "DIN 4109-32 Gl.(21), 100..720 kg/m2"),
        ("Eingebaut Trittschall", "L'_n,w = L_n,eq,0,w - dL_w + K_T", "DIN 4109-2 Gl.(22); dL_w aus DIN 4109-34"),
        ("Nachweis Luftschall", "vorh. R'_w  >=  erf. R'_w", "erfuellt, wenn wahr"),
        ("Nachweis Trittschall", "vorh. L'_n,w  <=  zul. L'_n,w", "erfuellt, wenn wahr"),
    ]
    r = 3
    r = _hdr(ws, r, ["Groesse", "Formel", "Quelle / Gueltigkeit"])
    for name, f, q in formeln:
        ws.cell(r, 1, name).font = _LBL
        fc = ws.cell(r, 2, f)
        fc.font = _MONO
        ws.cell(r, 3, q)
        for c in range(1, 4):
            ws.cell(r, c).border = _BORDER
        r += 1
    r += 1
    ws.cell(r, 1, "Anforderungswerte DIN 4109-1 (verwendete Rollen)").font = _BLOCK
    ws.cell(r, 1).fill = _BLOCKFILL
    r += 1
    r = _hdr(ws, r, ["DIN-Rolle", "Tabelle/Zeile", "erf. R'w [dB]", "zul. L'n,w [dB]", "Bezeichnung"])
    used_rollen = {z.din_rolle for z in res.zeilen if z.din_rolle}
    reqs = requirements()
    for key in [k for k in reqs if k in used_rollen]:
        rq = reqs[key]
        ws.cell(r, 1, key)
        ws.cell(r, 2, f"Tab.{rq['tabelle']} Z.{rq['zeile']}")
        ws.cell(r, 3, rq["erf_Rw"] if rq["erf_Rw"] is not None else "-").alignment = _RIGHT
        ws.cell(r, 4, rq["zul_Lnw"] if rq["zul_Lnw"] is not None else "-").alignment = _RIGHT
        ws.cell(r, 5, rq["bezeichnung"])
        r += 1
    _widths(ws, {1: 26, 2: 14, 3: 14, 4: 16, 5: 52})


# ========================================================== Blatt 4: Rechengang (live)
def _blatt_rechengang(wb, res: NachweisErgebnis):
    ws = wb.create_sheet("4 Rechengang")
    ws["A1"] = "Rechengang je Bauteil  —  Zellen enthalten LIVE-Excel-Formeln (aenderbar)"
    ws["A1"].font = _TITLE
    r = 3
    for z in res.zeilen:
        r = _rechengang_bauteil(ws, r, z, res.projekt)
        r += 1
    _widths(ws, {1: 34, 2: 14, 3: 14, 4: 16, 5: 46})


def _scalar(ws, r, label, value, *, fmt=None, formula=False, source="", bold=False):
    ws.cell(r, 1, label).font = _LBL if bold else Font()
    c = ws.cell(r, 4, value)
    c.alignment = _RIGHT
    if formula:
        c.font = _MONO
    if fmt:
        c.number_format = fmt
    if source:
        ws.cell(r, 5, source)
    return r + 1


def _rechengang_bauteil(ws, r, z: NachweisZeile, projekt) -> int:
    e = z.ergebnis
    # Blockkopf
    ws.cell(r, 1, f"{z.bauteil_id} — {z.bezeichnung}  ({z.typ})").font = _BLOCK
    for c in range(1, 6):
        ws.cell(r, c).fill = _BLOCKFILL
    r += 1

    # Tuer/Fenster: direkter Elementwert
    if not e.schichten and e.rw_konstruktion is not None:
        rw_row = _scalar(ws, r, "R_w Element (Datenblatt)", e.rw_konstruktion, fmt="0.0", source="Tuer/Fenster: Anforderung gilt fuer R_w")
        r = rw_row
        erf_row = r
        r = _scalar(ws, r, "erforderlich R_w", z.erf_rw if z.erf_rw is not None else "-", source=z.tabelle_zeile)
        nf = f'=IF(D{rw_row-1}>=D{erf_row},"erfuellt","NICHT erfuellt")'
        cell = ws.cell(r, 1, "Nachweis"); cell.font = _LBL
        nc = ws.cell(r, 4, nf); nc.font = _RESULT; nc.fill = _STATUSFILL[z.status]
        return r + 1

    if e.masse_kg_m2 is None:
        ws.cell(r, 1, "R_w nicht bestimmbar (kein Massivaufbau / kein Elementwert).")
        return r + 1

    # --- m' aus Schichten (live) ---
    r = _hdr(ws, r, ["Schicht / Material", "d [m]", "rho [kg/m3]", "m'_i = d*rho", "zaehlt zu m'?"])
    first = r
    for s in e.schichten:
        ws.cell(r, 1, s.material + (f"  [{s.erkannt_als}]" if s.erkannt_als else ""))
        if s.rohdichte is None:
            ws.cell(r, 4, "-")
            r += 1
            continue
        ws.cell(r, 2, round(s.dicke_mm / 1000.0, 4)).alignment = _RIGHT
        ws.cell(r, 3, s.rohdichte).alignment = _RIGHT
        mc = ws.cell(r, 4, f"=B{r}*C{r}"); mc.font = _MONO; mc.number_format = "0.0"
        ws.cell(r, 5, "x" if s.in_masse else "").alignment = _RIGHT
        r += 1
    last = r - 1
    # m' gesamt = Summe nur der zaehlenden Schichten
    mrow = r
    ws.cell(r, 1, "m' (flaechenbezogene Masse Rohbauteil)").font = _LBL
    mc = ws.cell(r, 4, f'=SUMIF(E{first}:E{last},"x",D{first}:D{last})')
    mc.font = _MONO; mc.number_format = "0.0"
    ws.cell(r, 5, "DIN 4109-32 Gl.(3)")
    mcell = f"D{mrow}"
    r += 1
    if e.schwimmender_estrich:
        ws.cell(r, 1, "Hinweis: schwimmender Estrich zaehlt nicht zur Rohdecke (DIN 4109-32, 4.8.4.2)")
        ws.cell(r, 1).font = Font(italic=True, color="9a6700")
        r += 1

    # --- Luftschall ---
    ws.cell(r, 1, "Luftschalldaemmung").font = _BLOCK
    r += 1
    rw_formel, rw_quelle = _rw_formel(e.massekurve, mcell)
    rw_row = r
    ws.cell(r, 1, "R_w (massiv) aus m'").font = _LBL
    c = ws.cell(r, 4, rw_formel); c.font = _MONO; c.number_format = "0.0"
    ws.cell(r, 5, rw_quelle)
    r += 1
    drw_row = r
    bt = next((b for b in projekt.bauteile if b.id == z.bauteil_id), None)
    r = _scalar(ws, r, "+ dR_w Vorsatzschale", bt.delta_rw_vorsatz if bt else 0, fmt="0.0", source="DIN 4109-34")
    k_row = r
    r = _scalar(ws, r, "- K Flankenuebertragung", e.k_luft or 0, fmt="0.0", source="DIN 4109-2 (vereinfacht)")
    rstrich_row = r
    ws.cell(r, 1, "R'_w (eingebaut)").font = _LBL
    c = ws.cell(r, 4, f"=D{rw_row}+D{drw_row}-D{k_row}"); c.font = _MONO; c.number_format = "0.0"
    ws.cell(r, 5, "R'w = R_w + dR_w - K")
    r += 1
    erf_row = r
    r = _scalar(ws, r, "erforderlich R'_w", z.erf_rw if z.erf_rw is not None else "-", source=z.tabelle_zeile, bold=True)
    luft_check_row = r
    ws.cell(r, 1, "Nachweis Luftschall").font = _LBL
    if z.erf_rw is not None:
        nc = ws.cell(r, 4, f'=IF(D{rstrich_row}>=D{erf_row},"erfuellt","NICHT erfuellt")')
        nc.font = _RESULT
        nc.fill = _GREEN if z.rw_ok else _RED
    else:
        ws.cell(r, 4, "keine Anforderung")
    r += 1

    # --- Trittschall (nur Decken) ---
    if e.lnw_strich is not None:
        ws.cell(r, 1, "Trittschalldaemmung").font = _BLOCK
        r += 1
        ln_row = r
        ws.cell(r, 1, "L_n,eq,0,w aus m'").font = _LBL
        c = ws.cell(r, 4, f"=164-35*LOG10({mcell})"); c.font = _MONO; c.number_format = "0.0"
        ws.cell(r, 5, "DIN 4109-32 Gl.(21)")
        r += 1
        dlw_row = r
        r = _scalar(ws, r, "- dL_w (Estrich/Belag)", e.dlw_verwendet or 0, fmt="0.0", source="DIN 4109-34")
        kt_row = r
        kt_src = f"DIN 4109-2 Gl.(26): 0,6+5,5*lg(m's/m'f,m), m'f,m={e.m_flank_mittel:.0f}" if e.m_flank_mittel else "DIN 4109-2"
        r = _scalar(ws, r, "+ K Flanke (Trittschall)", e.k_tritt or 0, fmt="0.0", source=kt_src)
        lnstrich_row = r
        ws.cell(r, 1, "L'_n,w (eingebaut)").font = _LBL
        c = ws.cell(r, 4, f"=D{ln_row}-D{dlw_row}+D{kt_row}"); c.font = _MONO; c.number_format = "0.0"
        ws.cell(r, 5, "L'n,w = L_n,eq - dL_w + K_T")
        r += 1
        zul_row = r
        r = _scalar(ws, r, "zulaessig L'_n,w", z.zul_lnw if z.zul_lnw is not None else "-", source=z.tabelle_zeile, bold=True)
        ws.cell(r, 1, "Nachweis Trittschall").font = _LBL
        if z.zul_lnw is not None:
            nc = ws.cell(r, 4, f'=IF(D{lnstrich_row}<=D{zul_row},"erfuellt","NICHT erfuellt")')
            nc.font = _RESULT
            nc.fill = _GREEN if z.lnw_ok else _RED
        r += 1
    return r


def _widths(ws, mapping: dict[int, int]):
    for col, w in mapping.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def export_rechengang_excel(res: NachweisErgebnis, pfad: str | Path) -> Path:
    pfad = Path(pfad)
    wb = Workbook()
    _blatt_uebersicht(wb, res)
    _blatt_eingabe(wb, res)
    _blatt_formeln(wb, res)
    _blatt_rechengang(wb, res)
    wb.save(pfad)
    return pfad
