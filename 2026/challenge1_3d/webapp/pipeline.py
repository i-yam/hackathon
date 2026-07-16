#!/usr/bin/env python3
"""Opening-detection pipeline (importable). Detect slab/wall openings from
construction-plan PDFs, dedup overlapping plans, build a formatted Excel
workbook, and return structured rows + plan preview PNGs."""
import fitz, re, math, os, base64
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCALE=50.0; PT_TO_MM=25.4/72.0*SCALE; DENSITY=2.4; DEFAULT_SLAB_CM=30.0; ATTACH_RADIUS_PT=40
PREFIX={'DDB':'Ceiling','DB':'Ceiling','WDB':'Wall','WD':'Wall'}
DIM_RE=re.compile(r'^Ø\s*(\d+(?:\.\d+)?)$|^(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)$')
JOIN_RE=re.compile(r'^(DDB|WDB|DB|WD)\s*(?:Ø\s*(\d+(?:\.\d+)?)|(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?))$')
INTRA_TOL_PT=6; MATCH_TOL_MM=300; OFFSET_BIN_MM=300; MIN_INLIERS=3

def _dim(t):
    t=t.replace(' ','');m=DIM_RE.match(t)
    if not m:return None
    if m.group(1):d=float(m.group(1));return('round',d,d)
    return('rectangular',float(m.group(2)),float(m.group(3)))
def _wt(L,W,H,g):
    a=(math.pi/4*L*W) if g=='round' else L*W
    return a*H*DENSITY/1000.0
def _levels(words):
    lv=[]
    for i,w in enumerate(words):
        m=re.search(r'RD(OK|UK)\s*=?\s*(-?\d+\.\d+)',w[4])
        if m:lv.append({'k':m.group(1),'v':float(m.group(2)),'x':(w[0]+w[2])/2,'y':(w[1]+w[3])/2});continue
        t=w[4].rstrip('=')
        if t in('RDOK','RDUK'):
            k='OK' if t.endswith('OK') else 'UK'
            for j in range(i+1,min(i+3,len(words))):
                mv=re.fullmatch(r'-?\d+\.\d+',words[j][4])
                if mv:lv.append({'k':k,'v':float(mv.group()),'x':(w[0]+w[2])/2,'y':(w[1]+w[3])/2});break
    return lv
def _slab(x,y,lv):
    ok=[l for l in lv if l['k']=='OK'];uk=[l for l in lv if l['k']=='UK']
    if not ok or not uk:return DEFAULT_SLAB_CM,None
    o=min(ok,key=lambda l:(l['x']-x)**2+(l['y']-y)**2);u=min(uk,key=lambda l:(l['x']-x)**2+(l['y']-y)**2)
    t=abs(o['v']-u['v'])*100.0
    return (round(t),o['v']) if 8<=t<=120 else (DEFAULT_SLAB_CM,o['v'])
def _fmt(v):return int(v) if float(v)%1==0 else round(v,1)

def parse_plan(path, render_zoom=0.5):
    d=fitz.open(path);p=d[0];words=p.get_text('words')
    name=os.path.splitext(os.path.basename(path))[0]
    floor=name.split('_')[1] if '_' in name and len(name.split('_'))>1 else 'NA'
    lv=_levels(words)
    dims=[{'geom':r[0],'L':r[1],'W':r[2],'x':(w[0]+w[2])/2,'y':(w[1]+w[3])/2,'used':False}
          for w in words for r in [_dim(w[4])] if r]
    raw=[]
    def emit(pre,g,L,W,cx,cy):
        typ=PREFIX[pre];H,lvl=_slab(cx,cy,lv) if typ=='Ceiling' else (DEFAULT_SLAB_CM,None)
        lab=('%sØ%g'%(pre,L)) if g=='round' else ('%s%g/%g'%(pre,L,W))
        raw.append({'plan':name,'floor':floor,'type':typ,'geometry':g,'length_cm':_fmt(L),
            'width_cm':_fmt(W),'diameter_cm':(_fmt(L) if g=='round' else None),'height_cm':H,
            'label':lab,'weight_kg':round(_wt(L,W,H,g),1),'slab_top_m':lvl,
            'x_pt':round(cx,1),'y_pt':round(cy,1),'x_mm':round(cx*PT_TO_MM),'y_mm':round(cy*PT_TO_MM)})
    for w in words:
        m=JOIN_RE.match(w[4].replace(' ',''))
        if not m:continue
        pre=m.group(1)
        if m.group(2):g,L,W='round',float(m.group(2)),float(m.group(2))
        else:g,L,W='rectangular',float(m.group(3)),float(m.group(4))
        emit(pre,g,L,W,(w[0]+w[2])/2,(w[1]+w[3])/2)
    for w in words:
        if w[4] not in PREFIX:continue
        cx,cy=(w[0]+w[2])/2,(w[1]+w[3])/2;best=None;bd=ATTACH_RADIUS_PT
        for dm in dims:
            if dm['used']:continue
            dist=math.hypot(dm['x']-cx,dm['y']-cy)
            pref=0 if(-32<dm['y']-cy<-8 and abs(dm['x']-cx)<14) else 8
            if dist+pref<bd:bd=dist+pref;best=dm
        if best is None:continue
        best['used']=True;emit(w[4],best['geom'],best['L'],best['W'],cx,cy)
    pix=p.get_pixmap(matrix=fitz.Matrix(render_zoom,render_zoom))
    png=base64.b64encode(pix.tobytes('png')).decode()
    meta={'name':name,'w':pix.width,'h':pix.height,'zoom':render_zoom,'png':png}
    d.close()
    return raw, meta

def process(paths):
    raw=[];metas=[]
    for pth in paths:
        ops,meta=parse_plan(pth);raw+=ops;metas.append(meta)
    byplan=defaultdict(list)
    for o in raw:byplan[o['plan']].append(o)
    clean=[]
    for plan,lst in byplan.items():
        kept=[]
        for o in lst:
            if any(o['label']==k['label'] and abs(o['x_pt']-k['x_pt'])<INTRA_TOL_PT
                   and abs(o['y_pt']-k['y_pt'])<INTRA_TOL_PT for k in kept):continue
            kept.append(o)
        clean+=kept
    for i,o in enumerate(clean):o['uid']=i
    plans=sorted({o['plan'] for o in clean});by=defaultdict(list)
    for o in clean:by[o['plan']].append(o)
    parent=list(range(len(clean)))
    def find(i):
        while parent[i]!=i:parent[i]=parent[parent[i]];i=parent[i]
        return i
    def union(i,j):
        ri,rj=find(i),find(j)
        if ri!=rj:parent[ri]=rj
    def register(A,B):
        votes=Counter();grp=defaultdict(list)
        for a in A:
            for b in B:
                if a['label']!=b['label']:continue
                dx=a['x_mm']-b['x_mm'];dy=a['y_mm']-b['y_mm']
                votes[(round(dx/OFFSET_BIN_MM),round(dy/OFFSET_BIN_MM))]+=1
                grp[(round(dx/OFFSET_BIN_MM),round(dy/OFFSET_BIN_MM))].append((dx,dy))
        if not votes:return None
        best,n=votes.most_common(1)[0]
        if n<MIN_INLIERS:return None
        g=grp[best];return sum(p[0] for p in g)/len(g),sum(p[1] for p in g)/len(g),n
    regs=0
    for i in range(len(plans)):
        for j in range(i+1,len(plans)):
            r=register(by[plans[i]],by[plans[j]])
            if not r:continue
            dx,dy,n=r;regs+=1
            for a in by[plans[i]]:
                ax,ay=a['x_mm']-dx,a['y_mm']-dy;best=None;bd=MATCH_TOL_MM
                for b in by[plans[j]]:
                    if a['label']!=b['label']:continue
                    dd=math.hypot(ax-b['x_mm'],ay-b['y_mm'])
                    if dd<bd:bd=dd;best=b
                if best is not None:union(a['uid'],best['uid'])
    clusters=defaultdict(list)
    for o in clean:clusters[find(o['uid'])].append(o)
    def sk(it):
        g=it[1][0];return(0 if g['type']=='Ceiling' else 1,g['plan'],g['y_pt'])
    master=[]
    for k,item in enumerate(sorted(clusters.items(),key=sk),start=1):
        grp=item[1];rep=grp[0];seen=sorted({g['plan'] for g in grp});oid='%03d'%k
        for idx,g in enumerate(grp):g['opening_id']=oid;g['is_duplicate']=(idx>0);g['n_plans']=len(seen)
        master.append({'opening_id':oid,'floor':rep['floor'],
            'plans':', '.join(seen),'appears_in_n_plans':len(seen),'label':rep['label'],
            'type':rep['type'],'geometry':rep['geometry'],'length_cm':rep['length_cm'],
            'width_cm':rep['width_cm'],'diameter_cm':rep.get('diameter_cm'),'height_cm':rep['height_cm'],
            'slab_top_m':rep['slab_top_m'],'weight_kg':rep['weight_kg'],'pos_x_mm':rep['x_mm'],'pos_y_mm':rep['y_mm']})
    return {'master':master,'clean':clean,'metas':metas,'plan_pairs_registered':regs}

# ---------- Excel ----------
RED='00C81E1E';DARK='001A1A1A';LIGHT='00F2F2F2';GREY='00777777'
def _hdr(c):
    th=Side(style='thin',color='00D0D0D0')
    c.font=Font(name='Arial',bold=True,color='00FFFFFF',size=10)
    c.fill=PatternFill('solid',fgColor=DARK)
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    c.border=Border(th,th,th,th)
def build_workbook(result, path):
    master=result['master'];clean=result['clean']
    slab=[m for m in master if m['type']=='Ceiling'];wall=[m for m in master if m['type']=='Wall']
    th=Side(style='thin',color='00D0D0D0');bd=Border(th,th,th,th)
    wb=Workbook()
    def sheet(ws,rows,title,sub):
        ws.sheet_view.showGridLines=False
        ws.cell(1,1,title).font=Font(name='Arial',bold=True,size=14,color=RED)
        ws.cell(2,1,sub).font=Font(name='Arial',italic=True,size=9,color=GREY)
        cols=['Opening ID','Floor','Plan(s)','# Plans','Label','Type','Geometry',
              'Length (cm)','Width (cm)','Diameter (cm)','Height (cm)','Slab top (m)','Weight (kg)','Pos X (mm)','Pos Y (mm)']
        for c,h in enumerate(cols,1):_hdr(ws.cell(4,c,h))
        Ln=get_column_letter(8);Wd=get_column_letter(9);Di=get_column_letter(10);Hg=get_column_letter(11)
        G=get_column_letter(7);WT=get_column_letter(13)
        for i,m in enumerate(rows):
            r=5+i; rnd=(m['geometry']=='round')
            vals=[m['opening_id'],m['floor'],m['plans'],m['appears_in_n_plans'],m['label'],m['type'],m['geometry'],
                  (None if rnd else m['length_cm']),(None if rnd else m['width_cm']),
                  ((m.get('diameter_cm', m['length_cm'])) if rnd else None),m['height_cm'],m['slab_top_m'],None,m['pos_x_mm'],m['pos_y_mm']]
            for c,v in enumerate(vals,1):
                cell=ws.cell(r,c,v);cell.border=bd;cell.font=Font(name='Arial',size=10)
                cell.alignment=Alignment(horizontal='left' if c==3 else 'center')
            f='=IF(%s%d="round",PI()/4*%s%d^2,%s%d*%s%d)*%s%d*2.4/1000'%(G,r,Di,r,Ln,r,Wd,r,Hg,r)
            wt=ws.cell(r,13,f);wt.number_format='0.0';wt.border=bd
            wt.font=Font(name='Arial',size=10,bold=True);wt.alignment=Alignment(horizontal='center')
            if m['appears_in_n_plans']>1:ws.cell(r,4).fill=PatternFill('solid',fgColor='00FFF2CC')
            elif i%2:
                for c in range(1,16):
                    if c!=4:ws.cell(r,c).fill=PatternFill('solid',fgColor=LIGHT)
        tr=5+len(rows);ws.cell(tr,1,'TOTAL').font=Font(name='Arial',bold=True)
        t=ws.cell(tr,13,'=SUM(%s5:%s%d)'%(WT,WT,tr-1));t.number_format='0.0';t.font=Font(name='Arial',bold=True,color=RED)
        for c,wd in enumerate([10,7,16,7,13,9,12,11,11,12,11,11,11,11,11],1):
            ws.column_dimensions[get_column_letter(c)].width=wd
        ws.freeze_panes='A5'
    ws1=wb.active;ws1.title='Slab Openings'
    sheet(ws1,slab,'Floor-Slab Openings (3D-print fabrication target)',
          'Deckendurchbruch (DDB) - scale 1:50 - weight = displaced concrete x 2.4 t/m3 - round = pipe penetration (diameter)')
    if wall:
        sheet(wb.create_sheet('Wall Openings'),wall,'Wall Openings (detected, separate trade)',
              'Wanddurchbruch (WDB) - detected for completeness')
    wsS=wb.create_sheet('Summary');wsS.sheet_view.showGridLines=False
    wsS.cell(1,1,'Summary').font=Font(name='Arial',bold=True,size=14,color=RED)
    rr=sum(1 for m in slab if m['geometry']=='round')
    rows=[('Plans processed',len(result['metas'])),('Raw detections',len(clean)),
          ('Unique openings (after dedup)',len(master)),
          ('Duplicates removed',len(clean)-len(master)),('',''),
          ('Slab openings (DDB)',len(slab)),('  round (diameter)',rr),('  rectangular',len(slab)-rr),
          ('  total weight (kg)',round(sum(m['weight_kg'] for m in slab),1)),('',''),
          ('Wall openings (WDB)',len(wall))]
    for i,(k,v) in enumerate(rows):
        wsS.cell(3+i,1,k).font=Font(name='Arial',size=11,bold=(k!='' and not k.startswith(' ') and not k[0].islower()))
        b=wsS.cell(3+i,2,v);b.font=Font(name='Arial',size=11,bold=True);b.alignment=Alignment(horizontal='right')
    wsS.column_dimensions['A'].width=34;wsS.column_dimensions['B'].width=14
    wb.save(path);return path
