#!/usr/bin/env python3
"""Build fabrication-ready opening list: Excel workbook (live formulas) + CSV.
Sheet 1 = slab openings (3D-print fabrication target); Sheet 2 = wall openings;
Sheet 3 = all raw detections (audit); Sheet 4 = summary."""
import json, os, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA=os.path.dirname(os.path.abspath(__file__))
d=json.load(open(os.path.join(DATA,'_openings_processed.json')))
master=d['master']; clean=d['clean']
slab=[m for m in master if m['type']=='Ceiling']
wall=[m for m in master if m['type']=='Wall']

RED='00C81E1E'; DARK='001A1A1A'; LIGHT='00F2F2F2'; GREY='00777777'
thin=Side(style='thin',color='00D0D0D0'); border=Border(thin,thin,thin,thin)
def hdr(c):
    c.font=Font(name='Arial',bold=True,color='00FFFFFF',size=10)
    c.fill=PatternFill('solid',fgColor=DARK)
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=border

wb=Workbook()
def opening_sheet(ws,rows,title,subtitle):
    ws.sheet_view.showGridLines=False
    t=ws.cell(1,1,title); t.font=Font(name='Arial',bold=True,size=14,color=RED)
    ws.cell(2,1,subtitle).font=Font(name='Arial',italic=True,size=9,color=GREY)
    cols=['Opening ID','Floor','Plan(s)','# Plans','Label','Type','Geometry',
          'Length (cm)','Width (cm)','Diameter (cm)','Slab height (cm)','Slab top (m)',
          'Weight (kg)','Pos X (mm)','Pos Y (mm)']
    r0=4
    for c,h in enumerate(cols,1): hdr(ws.cell(r0,c,h))
    L=get_column_letter(8); W=get_column_letter(9); D=get_column_letter(10); H=get_column_letter(11)
    G=get_column_letter(7); WT=get_column_letter(13)
    for i,m in enumerate(rows):
        r=r0+1+i
        is_round=(m['geometry']=='round')
        length=(None if is_round else m['length_cm'])
        width=(None if is_round else m['width_cm'])
        dia=((m.get('diameter_cm', m['length_cm'])) if is_round else None)
        vals=[m['opening_id'],m['floor'],m['plans'],m['appears_in_n_plans'],m['label'],
              m['type'],m['geometry'],length,width,dia,m['height_cm'],
              m['slab_top_m'],None,m['pos_x_mm'],m['pos_y_mm']]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.border=border; cell.font=Font(name='Arial',size=10)
            cell.alignment=Alignment(horizontal='left' if c==3 else 'center')
        f='=IF(%s%d="round",PI()/4*%s%d^2,%s%d*%s%d)*%s%d*2.4/1000'%(G,r,D,r,L,r,W,r,H,r)
        wt=ws.cell(r,13,f); wt.number_format='0.0'; wt.border=border
        wt.font=Font(name='Arial',size=10,bold=True); wt.alignment=Alignment(horizontal='center')
        if m['appears_in_n_plans']>1: ws.cell(r,4).fill=PatternFill('solid',fgColor='00FFF2CC')
        elif i%2:
            for c in range(1,16):
                if c!=4: ws.cell(r,c).fill=PatternFill('solid',fgColor=LIGHT)
    tr=r0+1+len(rows)
    ws.cell(tr,1,'TOTAL').font=Font(name='Arial',bold=True)
    tot=ws.cell(tr,13,'=SUM(%s%d:%s%d)'%(WT,r0+1,WT,tr-1))
    tot.number_format='0.0'; tot.font=Font(name='Arial',bold=True,color=RED)
    for c,wd in enumerate([11,7,18,7,13,10,12,12,12,13,11,14,11,11,11],1):
        ws.column_dimensions[get_column_letter(c)].width=wd
    ws.freeze_panes='A5'

ws1=wb.active; ws1.title='Slab Openings (DDB)'
opening_sheet(ws1,slab,'Floor-Slab Openings  —  U1  (3D-print fabrication target)',
   'Deckendurchbruch (DDB) · scale 1:50 · weight = displaced concrete volume x 2.4 t/m3 · round = pipe penetration (Ø)')
ws2=wb.create_sheet('Wall Openings (WDB)')
opening_sheet(ws2,wall,'Wall Openings  —  U1  (detected, separate trade)',
   'Wanddurchbruch (WDB) · detected for completeness · not part of slab 3D-print scope')

# audit sheet
wsA=wb.create_sheet('All Detections')
wsA.sheet_view.showGridLines=False
wsA.cell(1,1,'All raw detections (before cross-plan dedup) — audit trail').font=Font(name='Arial',bold=True,size=12,color=DARK)
ac=['Opening ID','Duplicate?','Plan','Label','Type','Geometry','L/Ø (cm)','W (cm)','Slab h (cm)','Weight (kg)','X (pt)','Y (pt)']
for c,h in enumerate(ac,1): hdr(wsA.cell(3,c,h))
for i,o in enumerate(sorted(clean,key=lambda x:(x['type'],x['plan'],x['opening_id']))):
    r=4+i
    vals=[o['opening_id'],'DUP' if o.get('is_duplicate') else '',o['plan'],o['label'],o['type'],
          o['geometry'],o['length_cm'],o['width_cm'],o['height_cm'],o['weight_kg'],o['x_pt'],o['y_pt']]
    for c,v in enumerate(vals,1):
        cell=wsA.cell(r,c,v); cell.border=border; cell.font=Font(name='Arial',size=9)
        cell.alignment=Alignment(horizontal='center')
        if o.get('is_duplicate'): cell.fill=PatternFill('solid',fgColor='00FCE4E4')
for c,wd in enumerate([11,10,13,13,9,12,9,9,11,11,8,8],1):
    wsA.column_dimensions[get_column_letter(c)].width=wd
wsA.freeze_panes='A4'

# summary
wsS=wb.create_sheet('Summary'); wsS.sheet_view.showGridLines=False
wsS.cell(1,1,'Summary — Floor U1').font=Font(name='Arial',bold=True,size=14,color=RED)
nslab_r=sum(1 for m in slab if m['geometry']=='round')
rows=[('Plans processed',6),('Raw detections',len(clean)),
      ('Unique physical openings (after dedup)',len(master)),
      ('Cross-plan + double-draw duplicates removed',len(clean)-len(master)),('',''),
      ('SLAB openings (DDB) — fabrication target',len(slab)),
      ('  · round (pipe penetrations Ø)',nslab_r),
      ('  · rectangular',len(slab)-nslab_r),
      ('  · appearing in >1 plan (deduped)',sum(1 for m in slab if m['appears_in_n_plans']>1)),
      ('  · total fabrication weight (kg)',round(sum(m['weight_kg'] for m in slab),1)),
      ('  · heaviest element (kg)',round(max(m['weight_kg'] for m in slab),1)),
      ('  · lightest element (kg)',round(min(m['weight_kg'] for m in slab),1)),('',''),
      ('WALL openings (WDB) — detected, separate trade',len(wall))]
for i,(k,v) in enumerate(rows):
    r=3+i; a=wsS.cell(r,1,k); b=wsS.cell(r,2,v)
    a.font=Font(name='Arial',size=11,bold=(k!='' and not k.startswith(' ') and not k[0].islower()))
    b.font=Font(name='Arial',size=11,bold=True,color=DARK); b.alignment=Alignment(horizontal='right')
wsS.column_dimensions['A'].width=46; wsS.column_dimensions['B'].width=14

wb.save(os.path.join(DATA,'U1_Opening_Report.xlsx'))
with open(os.path.join(DATA,'U1_Slab_Openings.csv'),'w',newline='') as f:
    w=csv.writer(f)
    w.writerow(['opening_id','floor','plans','n_plans','label','type','geometry',
                'length_cm','width_cm','diameter_cm','slab_height_cm','slab_top_m','weight_kg','pos_x_mm','pos_y_mm'])
    for m in slab:
        is_round=(m['geometry']=='round')
        w.writerow([m['opening_id'],m['floor'],m['plans'],m['appears_in_n_plans'],m['label'],m['type'],
                    m['geometry'],(None if is_round else m['length_cm']),(None if is_round else m['width_cm']),
                    ((m.get('diameter_cm', m['length_cm'])) if is_round else None),m['height_cm'],m['slab_top_m'],
                    m['weight_kg'],m['pos_x_mm'],m['pos_y_mm']])
print('wrote xlsx + csv:',len(slab),'slab,',len(wall),'wall')
